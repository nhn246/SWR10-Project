import os
import sys
import re
import subprocess
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import win32com.client
import ctypes
import time
from mainframe_functions import *
import datetime
from openpyxl.styles import PatternFill



# Functions Used
def getfieldname(fieldnumber, h2s=None, fieldname=None):
    fieldname = ""
    h2s = ""
    screen = ""
    flimscreenpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "flim screens", f"{fieldnumber}.txt")
    if len(fieldnumber) > 8:
        fieldnumber = fieldnumber[:8]
    if (not os.path.exists(flimscreenpath)) or (_datetodayvalue(os.path.getmtime(flimscreenpath))+7 < _datetodayvalue()):
        wait()
        mainframetype(2, 2, "flim{enter}")
        screen = mainframegetscreen("FIELD INQUIRY MENU")
        mainframetype(2, 66, fieldnumber)
        mainframetype(7, 4, "s{enter}")
        screen = mainframegetscreen(" ")
        while not ("*** GENERAL FIELD INQUIRY ***" in screen or "NO DATA FOUND FOR THIS FIELD" in screen):
            screen = mainframegetscreen(" ")
        if "NO DATA FOUND FOR THIS FIELD" in screen:
            print(16, "", f"{fieldnumber} is not a valid field number")
            exit()
        with open(flimscreenpath, "w") as f:
            f.write(screen)
    with open(flimscreenpath, "r") as f:
        screen = f.read()
    fieldname = ""
    if "*** GENERAL FIELD INQUIRY ***" in screen:
        fieldname = mainframereadscreen(screen, 4, 23, 35)
        h2s = mainframereadscreen(screen, 14, 29, 35)
    return fieldname


def getcol(sheetobj, coltitle):
    c = 1
    found = 0
    lastcol = sheetobj.Cells.SpecialCells(11).Column
    findobject = sheetobj.Rows(1).Find(coltitle, None, None, None, None, 1)
    print (findobject.Column)
    return findobject.Column

def uppercase(str):
    str.upper()
    return str

def shortcutfieldnames(shortcutname):
    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "shortcut field names.txt"), "r") as f:
        fieldnumber = ""
        for line in f:
            temp = line.strip().split("\t")
            if temp[0] == shortcutname:
                fieldnumber = temp[1]
    return fieldnumber




# Script title
script_title = "SWR 10 validate data script 234623624746"

# Close the window with the script title
# ctypes.windll.user32.EnumWindows(ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.POINTER(ctypes.c_int), ctypes.POINTER(ctypes.c_int))(lambda hwnd, lParam: ctypes.windll.user32.PostMessageA(hwnd, 0x10, 0, 0) if ctypes.create_string_buffer(255), ctypes.windll.user32.GetWindowTextA(hwnd, ctypes.byref(text), 255), text.value.decode() == script_title else True), 0)


# Get the user name
username = os.getlogin()

# Find the Excel application object
# excel_app = win32com.client.Dispatch("Excel.Application")

# Kill all running Excel instances
subprocess.call("taskkill /f /im excel.exe")

workbook_name = ""
workbook_obj = None
found = False

while not found:
    try:
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = True
    except:
        # Display a message box to ask the user to open a SWR 10 data spreadsheet
        answer = ctypes.windll.user32.MessageBoxW(0, "Open a SWR 10 data spreadsheet. Click CANCEL to quit. Click TRY AGAIN to TERMINATE all Excel instances. Click CONTINUE to keep going.", "Information", 0x40 | 0x1)
        if answer == 2:
            sys.exit()
        if answer == 10:
            subprocess.call("taskkill /f /im excel.exe")
        continue

    for temp_workbook_obj in excel_app.Workbooks:
        if "SWR 10 data (" in temp_workbook_obj.Name:
            found = True
            workbook_name = temp_workbook_obj.Name
            workbook_obj = temp_workbook_obj
            break

    if not found:
        # Display a message box to ask the user to open a SWR 10 data spreadsheet
        answer = ctypes.windll.user32.MessageBoxW(0, "Open a SWR 10 data spreadsheet. Click CANCEL to quit. Click TRY AGAIN to TERMINATE all Excel instances. Click CONTINUE to keep going.", "Information", 0x40 | 0x1)
        if answer == 2:
            sys.exit()
        if answer == 10:
            subprocess.call("taskkill /f /im excel.exe")


######################
fieldnames = [None] * 100
fieldnumbers = [None] * 100
numfields = 0
api = ""
dp = ""
letterdate = ""
lease = ""
well = ""
letterdate_dayvalue = ""
approvaldate_dayvalue = ""
approvaldate_longdate = ""
letterdate_longdate = ""
fieldname = ""
h2s = ""
receivedate = ""
dpissueddate = ""
spuddate = ""
surfcasingdate = ""


sheetobj = workbook_obj.Worksheets("Sheet1")
workbook_obj.Activate()
sheetobj.Activate()

current_row = excel_app.ActiveCell.Row

sheetobj.Rows(current_row - 1).Interior.ColorIndex = -4142
sheetobj.Rows(current_row).Interior.ColorIndex = 4
workbook_name = excel_app.ActiveWorkbook.Name

first_row = excel_app.Selection.Row
last_row = 1
for row_object in excel_app.Selection.Rows:
    last_row = row_object.Row

for current_row in range(first_row, last_row + 1):
    if workbook_obj.Path != "S:\\PERMITTING\\ENG\\SWR 10\\SWR 10":
        ctypes.windll.user32.MessageBoxW(0, f'ERROR: the workbook file location is "{workbook_obj.Path}" instead of "S:\\PERMITTING\\ENG\\SWR 10\\SWR 10"', "Error", 0x10)
        sys.exit()

    sheetobj.Cells(current_row, getcol(sheetobj, "API#")).Activate()

    print(f"current row # {current_row}")

    attn = sheetobj.cell(current_row, getcol(sheetobj, "ATTN")).text

    api = sheetobj.cell(current_row, getcol(sheetobj, "API#")).value
    while len(api) != 9:
        sheetobj.cell(current_row, getcol(sheetobj, "API#")).activate()
        answer = input(f"enter API number\nexcel row # {current_row}")
        if answer == "2":
            sys.exit()
        api = sheetobj.cell(current_row, getcol(sheetobj, "API#")).value

    receivedate = sheetobj.cell(current_row, getcol(sheetobj, "receive date")).text
    while receivedate == "":
        sheetobj.cell(current_row, getcol(sheetobj, "receive date")).activate()
        answer = input(f"enter receive date\nexcel row # {current_row}")
        if answer == "2":
            sys.exit()
        receivedate = sheetobj.cell(current_row, getcol(sheetobj, "receive date")).text

    blanket = sheetobj.cell(current_row, getcol(sheetobj, "blanket (yes or no)")).value
    while "yes" not in blanket and "no" not in blanket:
        sheetobj.cell(current_row, getcol(sheetobj, "blanket (yes or no)")).activate()
        answer = input(f"blanket (yes or no)\nexcel row # {current_row}")
        if answer == "2":
            sys.exit()
        blanket = sheetobj.cell(current_row, getcol(sheetobj, "blanket (yes or no)")).text

    if blanket == "yes":
        sheetobj.cell(current_row, getcol(sheetobj, "Docket #")).value = ""

    if blanket == "no":
        TOCprod = sheetobj.cell(current_row, getcol(sheetobj, "production string TOC")).value
        while TOCprod == "":
            sheetobj.cell(current_row, getcol(sheetobj, "production string TOC")).activate()
            answer = input(f"must enter TOC if the application is non-blankets\nexcel row # {current_row}")
            if answer == "2":
                sys.exit()
            TOCprod = sheetobj.cell(current_row, getcol(sheetobj, "production string TOC")).value


time.sleep(1)


mainframetype(2, 2, "WBTM{enter}")
screen = mainframegetscreen("WELL BORE TECHNICAL DATA MENU")
mainframetype(8, 23, api[:3])
mainframetype(8, 27, api[4:9])
mainframetype(11, 8, "s")
mainframetype(14, 7, "s{enter}")

permithistoryscreen = mainframegetscreen("PERMIT NUMBERS AND WELLS WITHIN WELL BORE")
currentrowow = 5
foundpermit = 0
while True:
    currentrowow += 1
    permit = mainframereadscreen(permithistoryscreen, currentrowow, 7, 6)
    if len(permit) == 6 and permit.isdigit():
        foundpermit = 1
    if "___" in permit or len(permit) <= 1:
        break

if foundpermit == 1:
    mainframetype(currentrowow-1, 4, "s")
    mainframetype(18, 4, "s{enter}")

    permitscreen = mainframegetscreen("DRILLING PERMIT MASTER DATA INQUIRY")
    opname = mainframereadscreen(permitscreen, 5, 17, 34)
    lease = mainframereadscreen(permitscreen, 9, 17, 34)
    district = mainframereadscreen(permitscreen, 10, 17, 2)
    county = ""
    dpissueddate = mainframereadscreen(permitscreen, 15, 17, 10)
    dpissueddate = dpissueddate.replace(" ", "-")
    dp = mainframereadscreen(permitscreen, 3, 17, 6)
    spuddate = mainframereadscreen(permitscreen, 19, 42, 10)
    surfcasingdate = mainframereadscreen(permitscreen, 19, 17, 10)
    if api[0] == "6" or api[0] == "7":
        county = mainframereadscreen(permitscreen, 12, 17, 34)
    else:
        county = mainframereadscreen(permitscreen, 11, 17, 34)
    well = mainframereadscreen(permitscreen, 9, 69, 10)
    opnumber = mainframereadscreen(permitscreen, 5, 69, 6)
    api = mainframereadscreen(permitscreen, 12, 69, 3) + "-" + mainframereadscreen(permitscreen, 12, 73, 5)
    if "PERMIT TYPE => DRILL" in permitscreen:
        sheetobj.Cells(current_row, getcol(sheetobj, "new drill (yes/no)")).Value = "yes"
    else:
        sheetobj.Cells(current_row, getcol(sheetobj, "new drill (yes/no)")).Value = "no"
else:
    mainframetype(16, 4, "s@B@Bs{enter}")

    W2G1screen = mainframegetscreen("OIL AND GAS W-2/G-1 RECORD")
    opname = mainframereadscreen(W2G1screen, 5, 9, 34)
    lease = mainframereadscreen(W2G1screen, 4, 48, 31)
    api = mainframereadscreen(W2G1screen, 2, 9, 3) + "-" + mainframereadscreen(W2G1screen, 2, 15, 5)
    district = mainframereadscreen(W2G1screen, 3, 9, 2)
    county = mainframereadscreen(W2G1screen, 3, 66, 14)
    well = mainframereadscreen(W2G1screen, 3, 36, 9)

    time.sleep(2)

#check line 265
    mainframetype(10, 10, f"ornq " + {opname} + "s{enter}")

    screen = mainframegetscreen("ORGANIZATION NAME INQUIRY")
    opnumber = mainframereadscreen(screen, 4, 42, 6)

    ##################################

wait = print(1 + 262144, "", f"{opname}\n\nClick OK to change operator number", 2)
if wait == 1:
    while True:
        opnumber = input(f"Confirm operator number\n\n{opname}")
        if opnumber == "":
            sheetobj.rows[current_row].interior.colorindex = -4142
            break
        if opnumber.isdigit() and len(opnumber) == 6:
            break



line1 = [None] * 100
line2 = [None] * 100
line3 = [None] * 100
scriptdir = os.path.dirname(os.path.abspath(__file__))


ormq_screen_path = os.path.join(scriptdir, f"ormq screens{os.path.sep}{opnumber}.txt")
file_modified_date = os.path.getmtime(ormq_screen_path)
p5screen = ""
time_today = time.time()


if (not os.path.exists(ormq_screen_path)) or file_modified_date + 14 < time_today:    
# if (not os.path.exists(ormq_screen_path)) or (date_to_day_value(*file_modified_date) + 14 < date_to_day_value(year, mon, mday)):
    time.sleep(2)
    mainframetype(2, 2, f"ormq " + {opnumber} + "s{enter}")

    p5screen = mainframegetscreen("OPERATOR NUMBER")
    with open(ormq_screen_path, 'w') as output:
        output.write(p5screen)
else:
    with open(ormq_screen_path, 'r') as input_file:
        p5screen = input_file.read()

opnumber = mainframereadscreen(p5screen, 3, 19, 6)
opname = mainframereadscreen(p5screen, 4, 12, 34)
line1[0] = mainframereadscreen(p5screen, 6, 3, 38)
line2[0] = mainframereadscreen(p5screen, 7, 3, 38)
line3[0] = mainframereadscreen(p5screen, 8, 3, 38)

index = 0
if "@" not in attn:
    screen = ""
    time.sleep(2)
    mainframetype(2, 2, f"ORAR " + {opnumber} + "s{enter}")

    i = 1

    while True:
        screen = mainframegetscreen("ADDRESS INQUIRY")
        mainframetype(1, 1, "{f5}")
        time.sleep(100)
        screen = mainframegetscreen("ADDRESS INQUIRY")

        for r in range(4, 17, 4):
            for c in range(2, 45, 42):
                line1[i] = mainframereadscreen(screen, r, c, 29)
                line2[i] = mainframereadscreen(screen, r + 1, c, 29)
                line3[i] = mainframereadscreen(screen, r + 2, c, 29)
                i += 1
            if "NO MORE ADDRESSES" in screen:
                break
        num_addresses = i

        prompt = ""
        for i in range(num_addresses):
            prompt += f"{i}. {line1[i]} {line2[i]} {line3[i]}\n"
            if (i + 1) % 5 == 0:
                prompt += "\n"

        index = 0
        attn = attn.upper()
        index = input("Select Address", prompt, "0", "", 700, 700)
        if index == "":
            index = 0

##############################
op_address1 = line1[index]
op_address2 = line2[index]
op_address3 = line3[index]

if ("ATTN" in op_address1 or "C/O" in op_address1) and attn != "" and "@" not in attn:
    op_address1 = f"ATTN {attn}"

if op_address3 == "":
    op_address3 = op_address2
    op_address2 = op_address1
    op_address1 = "ATTN: REGULATORY DEPARTMENT"
    if attn != "" and "@" not in attn:
        op_address1 = f"ATTN {attn}"
        
if ("ATTN" in op_address1 or "C/O" in op_address1) and "@" in attn:
    op_address1 = "ATTN: REGULATORY DEPARTMENT"

sheetobj.cell(current_row, getcol(sheetobj, "OPERATOR ADDRESS 1")).value = op_address1
sheetobj.cell(current_row, getcol(sheetobj, "OPERATOR ADDRESS 2")).value = op_address2
sheetobj.cell(current_row, getcol(sheetobj, "OPERATOR ADDRESS 3")).value = op_address3

field_number_col = getcol(sheetobj, "field 1")
h2s_col = getcol(sheetobj, "h2s field 1")
fieldname_col = getcol(sheetobj, "field name 1")
field_perfs_col = getcol(sheetobj, "perfs field 1")
fieldname_display = ""
h2s_condition = "NO"
all_fields_letter_text = ""

if sheetobj.cell(current_row, field_number_col).text == "rsp":
    sheetobj.cell(current_row, field_number_col).value = 85280300
    sheetobj.cell(current_row, field_number_col + 1).value = 85280900
    sheetobj.cell(current_row, field_number_col + 2).value = 85448150

if sheetobj.cell(current_row, field_number_col).text == "oxy":
    sheetobj.cell(current_row, field_number_col).value = 85280300
    sheetobj.cell(current_row, field_number_col + 1).value = 56378750
    sheetobj.cell(current_row, field_number_col + 2).value = 55256030

if sheetobj.cell(current_row, field_number_col).text == "c":
    sheetobj.cell(current_row, field_number_col).value = 85280300
    sheetobj.cell(current_row, field_number_col + 1).value = 71021430
    sheetobj.cell(current_row, field_number_col + 2).value = 69765200
    sheetobj.cell(current_row, field_number_col + 3).value = 16559500

h2s=''
fieldname=''

while sheetobj.cell(current_row, field_number_col).value != "":
    field_number = str(sheetobj.cell(current_row, field_number_col).text).replace(" ", "")

    if not field_number.isdigit():
        field_number = shortcutfieldnames(field_number)

    if district == "7C" and sheetobj.cell(current_row, getcol(sheetobj, "field 1")).value == 85280300:
        sheetobj.cell(current_row, getcol(sheetobj, "field 1")).value = 85279200

    if district == "08" and sheetobj.cell(current_row, getcol(sheetobj, "field 1")).value == 85279200:
        sheetobj.cell(current_row, getcol(sheetobj, "field 1")).value = 85280300

    sheetobj.cell(current_row, field_number_col).value = field_number

    getfieldname(field_number, h2s, fieldname)
    if h2s == "PRESENT":
        h2s_condition = "YES"
    sheetobj.cell(current_row, fieldname_col).value = fieldname
    sheetobj.cell(current_row, h2s_col).value = h2s
    perfs = sheetobj.cell(current_row, field_perfs_col).value

    while len(perfs) < 2 and blanket == "no":
        sheetobj.cell(current_row, field_perfs_col).activate()
        print(1 + 262144, "", f"enter perfs\n\n{fieldname}")
        perfs = sheetobj.cell(current_row, field_perfs_col).value

    field_number = sheetobj.cell(current_row, field_number_col).value

    fieldname_display += f"{fieldname}{{{field_number}}}{{{perfs}}}\n"
    all_fields_letter_text += f\
    
    h2s_col += 1
    field_number_col += 1
    fieldname_col += 1
    field_perfs_col += 1

all_fields_letter_text = all_fields_letter_text.rstrip("; ")

if (sheetobj.cell(current_row, getcol(sheetobj, "field 1")).value == 85280300
        or sheetobj.cell(current_row, getcol(sheetobj, "field 1")).value == 85279200):
    sheetobj.cell(current_row, getcol(sheetobj, "Allow.")).value = "515 BOPD"

pos = len(all_fields_letter_text) - 1
while pos > 0:
    if all_fields_letter_text[pos:pos + 2] == "; ":
        all_fields_letter_text = all_fields_letter_text[:pos] + " and " + all_fields_letter_text[pos + 2:]
        break
    pos -= 1

field_name_display = field_name_display.rstrip(";")
sheetobj.cell(current_row, getcol(sheetobj, "Fields")).value = field_name_display
sheetobj.cell(current_row, getcol(sheetobj, "Field Assgn.")).value = sheetobj.cell(current_row, getcol(sheetobj, "field name 1")).value
    

today = datetime.date.today()

for current_row in range(1, sheetobj.max_row + 1):
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "disposition date")).value = today.strftime("%Y/%m/%d")
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "script run date")).value = today.strftime("%Y/%m/%d")
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "District")).value = district
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "County")).value = county
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "Operator")).value = opname
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "Operator no.")).value = opnumber
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "Lease")).value = lease
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "Well")).value = well
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "well name combo")).value = f"{lease} — WELL NO. {well}, API NO. {api}"
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "h2s restriction")).value = h2scondition
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "drilling permit issued date")).value = dpissueddate
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "drilling permit no.")).value = dp
    sheetobj.cell(row=current_row, column=getcol(sheetobj, "all fields letter body")).value = allfieldslettertext

    if spuddate != "" or surfcasingdate != "":
        sheetobj.cell(row=current_row, column=getcol(sheetobj, "current permit has spud date (yes/no)")).value = "yes"
    else:
        sheetobj.cell(row=current_row, column=getcol(sheetobj, "current permit has spud date (yes/no)")).value = "no"

    leavepermitopen = sheetobj.cell(row=current_row, column=getcol(sheetobj, "requested permit held open (yes/no)")).value
    newdrill = sheetobj.cell(row=current_row, column=getcol(sheetobj, "new drill (yes/no)")).value
    alreadyonschedule = sheetobj.cell(row=current_row, column=getcol(sheetobj, "all zones already on schedule (yes/no)")).value

    if blanket == "yes":
        sheetobj.cell(row=current_row, column=getcol(sheetobj, "requested permit held open (yes/no)")).value = "no"

    if sheetobj.cell(row=current_row, column=getcol(sheetobj, "new drill (yes/no)")).value != "yes" and blanket == "no":
        while sheetobj.cell(row=current_row, column=getcol(sheetobj, "all zones already on schedule (yes/no)")).value == "":
            # Replaceprint with an equivalent Python function, e.g., a simple print statement or a custom function
            print("Error: answer \"all zones already on schedule (yes/no)\"")

    sheetobj.cell(current_row, getcol(sheetobj, "Expiration Date")).formula = sheetobj.cell(current_row, getcol(sheetobj, "disposition date")).formula + 731
    sheetobj.cell(current_row, getcol(sheetobj, "expiration condition")).value = "This exception to SWR 10 will expire if not used within two (2) years from the date of this permit."

    if blanket == "no" and sheetobj.cell(current_row, getcol(sheetobj, "new drill (yes/no)")).value == "no" and sheetobj.cell(current_row, getcol(sheetobj, "all zones already on schedule (yes/no)")).value == "no":
        sheetobj.cell(current_row, getcol(sheetobj, "expiration condition")).value = f"This exception to SWR 10 will expire if not used within two (2) years from the original date of issuance for drilling permit no. {dp}."
        sheetobj.cell(current_row, getcol(sheetobj, "Expiration Date")).formula = sheetobj.cell(current_row, getcol(sheetobj, "drilling permit issued date")).formula + 731

    if sheetobj.cell(current_row, getcol(sheetobj, "requested permit held open (yes/no)")).value == "yes":
        sheetobj.cell(current_row, getcol(sheetobj, "expiration condition")).value = f"This exception to SWR 10 will expire if not used within two (2) years from the original date of issuance for drilling permit no. {dp}."
        sheetobj.cell(current_row, getcol(sheetobj, "Expiration Date")).formula = sheetobj.cell(current_row, getcol(sheetobj, "drilling permit issued date")).formula + 731
    

    permitconditions = ""
    if h2scondition == "YES":
        permitconditions += "The commingled well will be subject to Statewide Rule 36 (operation in hydrogen sulfide areas) because at least one of the commingled fields requires a Certificate of Compliance for Statewide Rule 36.  The well must be operated in accordance with Statewide Rule 36.\n\n"
    if blanket == "YES":
        permitconditions += "The completion report for the commingled well must indicate which perforations belong to which field.  The Commission may also require a wellbore diagram to be filed with the completion report for the commingled well.  If filed, the wellbore diagram must indicate which perforations belong to which field.\n\n"
    if blanket == "NO":
        permitconditions += "The completion of the commingled well must be a reasonable match with the wellbore diagram filed with the application.  Variances in completion depths are acceptable provided that these completion depths remain within the designated correlative intervals for the commingled fields.  A copy of this wellbore diagram must be filed with the completion report for the commingled well.\n\n"

    custompermitconditions = sheetobj.cell(row=current_row, column=getcol(sheetobj, "custom permit conditions")).value
    if custompermitconditions != "":
        permitconditions += custompermitconditions + "\n\n"

    if "@" in attn:
        permitconditions += f"Note: The distribution of this document will be by E-MAIL ONLY.  E-mail sent to {attn}.\n\n"

    if permitconditions != "":
        sheetobj.cell(row=current_row, column=getcol(sheetobj, "permit conditions")).value = "Permit conditions:\n\n" + permitconditions


    sheetobj.range(f"A{current_row}:A{current_row}").activate

    sheetobj.column(getcol(sheetobj, "ATTN")).hyperlinks.delete()

    sheetobj.cell(row=current_row, column=getcol(sheetobj, "validate run date")).value = today.strftime("%Y/%m/%d")

    sheetobj.cell(row=current_row, column=getcol(sheetobj, "disposition")).value = "pending"

    sheetobj.row_dimensions[current_row].fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

    workbook_obj.save() 



